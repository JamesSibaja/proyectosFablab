import numpy as np
from keras.models import Sequential
from keras.layers.core import Dense
import openpyxl
#import matplotlib.pyplot as plt
import tensorflow as tf
import tkinter as tk
from tkinter import ttk
import math


# doc = openpyxl.load_workbook ("datos_proyecto.xlsx", data_only=True)
# hoja = doc.get_sheet_by_name("datos")

doc2 = openpyxl.load_workbook ("datos_hidro1.xlsx", data_only=True)
hoja2 = doc2.get_sheet_by_name("dato2")

l=[]
# entrada=[]

# for row in hoja.iter_rows(min_row = 2,max_row = 20,min_col = 6,max_col = 16):
#     for cell in row:
#         dato = cell.value
#         if dato is None:
#             dato = 0
#         l.append(dato)
#     entrada.append(l)
#     l=[]

# salida=[]

# for row in hoja.iter_rows(min_row = 2,max_row = 20,min_col = 17):
#     for cell in row:
#         dato = cell.value
#         if dato is None:
#             dato = 0
#         l.append(dato)
#     salida.append(l)
#     l=[]

test=[]


for row in hoja2.iter_rows(min_row = 3,max_row = 70,min_col = 6,max_col = 16):
    for cell in row:
        dato = cell.value
        if dato is None:
            dato = 0
        l.append(dato)
    test.append(l)
    l=[]
#print(test)
testOut=[]

for row in hoja2.iter_rows(min_row = 3,max_row = 70,min_col = 17):
    for cell in row:
        dato = cell.value
        if dato is None:
            dato = 0
        l.append(dato)
    testOut.append(l)
    l=[]

#print(testOut)
 
# cargamos las 4 combinaciones de las compuertas XOR
training_data = np.array(test, "float32")
 
# y estos son los resultados que se obtienen, en el mismo orden
target_data = np.array(testOut, "float32")
 
model = Sequential()
model.add(Dense(128, input_dim=11, activation='relu'))
model.add(Dense(128, activation='relu'))
model.add(Dense(128, activation='relu'))
model.add(Dense(128, activation='relu'))
model.add(Dense(128, activation='relu'))
model.add(Dense(128, activation='relu'))

model.add(Dense(2))
 
model.compile(loss='mean_squared_error',
              optimizer=tf.keras.optimizers.Adam(0.0001))
 
historial=model.fit(training_data, target_data, epochs=25000, verbose=False)
 
def analisisDatos(ph,ce,act):
    agua = 0
    urea = 0
    phd = 0

    if act == 0:
        agua = 5
    if act == 1:
        urea = 50
    if act == 2:
        phd = 20
    
    resultado = model.predict([[300, 72, ce, ph, agua, 0, 0, 0, phd, urea, 108]])
    etiqueta_resultado.config(text=f"CE: {round(resultado[0][0],4)} pH:{round(resultado[0][1],4)}")
    print(act)

root = tk.Tk()
root.title("Modelo del hidropónico")
root.config(width=300, height=200)
etiqueta_ce = ttk.Label(text="CE:")
etiqueta_ph = ttk.Label(text="pH:")
etiqueta_ce.place(x=20, y=20)
etiqueta_ph.place(x=20, y=50)
# Crear caja de texto.
entry = ttk.Entry()
entry2 = ttk.Entry()
entry.place(x=50, y=20)
entry2.place(x=50, y=50)
combo = ttk.Combobox(
    state="readonly",
    values=["+5 gal", "+50 urea", "+20cc pHd","Ninguno"]
) 
button = ttk.Button(text="Predecir", command=lambda:  analisisDatos(ce=float(entry.get()),ph=float(entry2.get()),act=combo.current()))
button.place(x=50, y=110)
combo.place(x=50, y=80)
etiqueta_resultado = ttk.Label(text="")
etiqueta_resultado.place(x=20, y=140)
root.mainloop()




print("Hagamos una predicción!")
resultado = model.predict([[24.6666666666667, 1.5, 0.305, 8.1, 0, 0, 236.5, 94.6, 0, 0, 108]])

datoR = [1.27,7.1]
print("El resultado es " + str(resultado) )
print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/(datoR[0]-0.305))*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/(datoR[1]-8.1))*100),3)))

print("Hagamos una predicción!")
resultado = model.predict([[23.1666666666667, 23.1666666666667, 0.311, 7.75, 25, 0, 0, 0, 0, 0, 108]])

datoR = [0.305,8.1]
print("El resultado es " + str(resultado) )
print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/(datoR[0]-0.311))*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/(datoR[1]-7.75))*100),3)))

print("Hagamos una predicción!")
resultado = model.predict([[42.55, 17.833333333333, 1.27, 7.1, 0, 0, 0, 0, 0, 0, 108]])

datoR = [1.19,7.6]
print("El resultado es " + str(resultado) )
print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/(datoR[0]-1.27))*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/(datoR[1]-7.1))*100),3)))

print("\n Hagamos una predicción!")
resultado = model.predict([[24,5.3, 0.313,8.31, 5, 0, 283.5, 113.5, 0, 50, 108]])
datoR = [1.182,7.62]
print("El resultado es " + str(resultado) )
print("El real es " + str(datoR) )
print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/(datoR[0]-0.313))*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/(datoR[1]-8.31))*100),3)))

print("\n Hagamos una predicción!")
resultado = model.predict([[168,6, 0.94,7.95, 0, 0, 0, 0, 0, 50, 108]])
datoR = [1.181,7.76]
print("El resultado es " + str(resultado) )
print("El real es " + str(datoR) )
print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/(datoR[0]-0.94))*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/(datoR[1]-7.95))*100),3)))

print("\nHagamos una predicción!")
resultado = model.predict([[384,3.5, 1.902,8, 5, 0, 0, 0, 0, 0, 108]])
datoR = [1.75,7.89]
print("El resultado es " + str(resultado) )
print("El real es " + str(datoR) )
print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/(datoR[0]-1.902))*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/(datoR[1]-8))*100),3)))

print("\nHagamos una predicción!")
resultado = model.predict([[336,5.5, 1.83,8.1, 5, 0, 0, 0, 0, 50, 108]])
datoR = [1.539,7.87]
print("El resultado es " + str(resultado) )
print("El real es " + str(datoR) )
print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/(datoR[0]-1.83))*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/(datoR[1]-8.1))*100),3)))

# print("Hagamos una predicción!")
# resultado = model.predict([[336,5.5, 1.83,8.1, 5, 0, 0, 0, 0, 50, 108]])
# datoR = [1.539,7.87]
# print("El resultado es " + str(resultado) )
# print("El real es " + str(datoR) )
# print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/datoR[0])*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/datoR[1])*100),3)))

# print("Hagamos una predicción!")
# resultado = model.predict([[336,5.5, 1.83,8.1, 5, 0, 0, 0, 0, 50, 108]])
# datoR = [1.539,7.87]
# print("El resultado es " + str(resultado) )
# print("El real es " + str(datoR) )
# print("El error es " + str(round((abs((resultado[0][0]-datoR[0])/datoR[0])*100),3)) + " y " + str(round((abs((resultado[0][1]-datoR[1])/datoR[1])*100),3)))

#print("\n%s: %.2f%%" % (model.metrics_names[1], scores[1]*100))
#print (model.predict(training_data))